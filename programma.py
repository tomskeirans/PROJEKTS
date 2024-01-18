import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook 

komandas_nosaukums = input("Ievadiet komandas saīsinājumu: ").upper()

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

url = "https://www.floorball.lv/lv/2023/chempionats/v1/statistika_komandas#kopsavilkums"
driver.get(url)

wb=load_workbook('komandas.xlsx')
ws=wb.active
max_row=ws.max_row
for i in range(2,max_row+1):
    komandas_saisinajums = ws['B' + str(i)].value
    if komandas_saisinajums == komandas_nosaukums:
        komandas_pilnais = ws['A' + str(i)].value
    
xpath = f"//tr[contains(td[2], '{komandas_nosaukums}')]"
komandas_rinda = driver.find_element(By.XPATH, xpath)
soda_minutes = komandas_rinda.find_element(By.XPATH, "./td[24]").text
varti_vairakuma = komandas_rinda.find_element(By.XPATH, "./td[12]").text

vairakumu_skaits = round(int(soda_minutes)/2)
vairakuma_procents = round((int(varti_vairakuma)/int(vairakumu_skaits))*100)

print(f"{komandas_nosaukums} pilnais komandas nosaukums: {komandas_pilnais}")
print(f"Soda minūtes pret komandu: {soda_minutes}")
print(f"{komandas_nosaukums} vārti vairākumā: {varti_vairakuma}")
print(f"{komandas_nosaukums} vairākumi: {vairakumu_skaits}")
print(f"{komandas_nosaukums} vairākuma procents: {vairakuma_procents}%")

driver.quit()